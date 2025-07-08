from tkinter import *
from tkinter import ttk

# Sample product data
products = {
    0: ['a', 20],
    1: ['b', 10],
    2: ['c', 30],
    3: ['d', 15],
}

# Main Window
gem = Tk()
gem.geometry('1000x800')
gem.title("Tool Shop")

# Header Label
Label(gem, text='Web for Sell Tools', fg="gold", bg="black", width=145, height=2).pack()

# Treeview for order summary
tree_frame = Frame(gem)
tree_frame.pack(pady=10)
tree = ttk.Treeview(tree_frame, columns=("Name", "Qty", "Total"), show='headings', height=5)
tree.heading("Name", text="Item")
tree.heading("Qty", text="Qty")
tree.heading("Total", text="Total")
tree.pack()

# Spinboxes to select quantities
spinbox_vars = []
for i in range(len(products)):
    var = IntVar()
    Spinbox(gem, from_=0, to=10, textvariable=var, width=5).pack()
    spinbox_vars.append(var)

# Calculate & display bill
def calculate():
    for item in tree.get_children():
        tree.delete(item)
    total = 0
    for i, var in enumerate(spinbox_vars):
        qty = var.get()
        if qty > 0:
            price = products[i][1]
            subtotal = qty * price
            total += subtotal
            tree.insert("", "end", values=(products[i][0], qty, subtotal))
    Label(gem, text=f"Total: {total}", bg="white").pack()

# Buy Button
Button(gem, text="Buy", command=calculate, bg="gray", fg="white").pack(pady=20)

gem.mainloop()
def fatora():
    global p
    tree.delete(*tree.get_children())  # مسح الجدول السابق
    total_all = 0

    for i in range(len(products)):
        qty = int(spinboxes[i].get())
        if qty > 0:
            name, price = products[i]
            total = qty * price
            po=qty
            total_all += total
            t=total_all+price
            tree.insert("", "end", values=(name, price, qty, total))

    p=bo_Entry.config(text=f"الإجمالي الكلي: {total_all} ريال")
    
    
    o=datetime.now()
    ti=o.strftime('%d -%m-%y')
    bo_Entry.insert("1",f"total={total_all}")
    co_Entry.insert("2",str(ti))
    pass
from flask import Flask, render_template, request, jsonify
from datetime import datetime
from openpyxl import Workbook

app = Flask(__name__)

products = {
    0: ["saw", 10],
    1: ["drill", 30],
    2: ["hammer", 15]
}

@app.route("/")
def index():
    return render_template("index.html", products=products)

@app.route("/generate-bill", methods=["POST"])
def generate_bill():
    data = request.json
    quantities = data["quantities"]
    name = data.get("name", "")
    phone = data.get("phone", "")
    location = data.get("location", "")
    bill = []
    total_all = 0

    for i, qty in enumerate(quantities):
        qty = int(qty)
        if qty > 0:
            product_name, price = products[i]
            total = qty * price
            bill.append({
                "name": product_name,
                "price": price,
                "qty": qty,
                "total": total
            })
            total_all += total
            print(total_all)




    return jsonify({
        "bill": bill,
        "total": total_all,
        
        
        
        "date": datetime.now().strftime("%d-%m-%Y"),
        "name": name,
        "phone": phone,
        "location": location
    })
w=Workbook()
ws=w.active
ws.title='custmur'
ws['A1']='name'
ws['B1']='phone'
ws['C1']='loction'
ws['D1']='total'
ws['E1']='price'
ws['F1']='tool'
w.save('ALI.xlsx')

if __name__ == "__main__":
    app.run(debug=True)


            ws=wb.active
            ws.title='custmur'
            ws['A1']='name'
            ws['B1']='phone'
            ws['C1']='loction'
            ws['D1']='total'
            ws['E1']='price'
            ws['F1']='tool'
            wb.save('ALI(2).xlsx')
            pass





